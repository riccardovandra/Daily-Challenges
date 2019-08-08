from openpyxl import Workbook

wb = Workbook()
ws = wb.active
tax_brackets = [10000, 30000, 100000]
marginal_tax_rate = [0.1, 0.25, 0.4]

for row in range(1, 4):  # create a set of numbers from 1 to 3
    ws.cell(row=row, column=1, value=tax_brackets[row-1])
    ws.cell(row=row, column=2, value=marginal_tax_rate[row-1])

wb.save('tax.xlsx')


# Tax Brackets
taxbracket1 = ws['A1'].value
taxbracket2 = ws['A2'].value
taxbracket3 = ws['A3'].value

# Marginal Tax  Rate
mtr1 = ws['B1'].value
mrt2 = ws['B2'].value
mrt3 = ws['B3'].value


def tax(income):
    if income < taxbracket1:
        print(0)
    elif income < taxbracket2:
        print((income-taxbracket1)*mtr1)
    elif income < taxbracket3:
        print((taxbracket2-taxbracket1)*mtr1+(income-taxbracket2)*mrt2)
    elif income > taxbracket3:
        print((taxbracket2-taxbracket1)*mtr1 +
              (taxbracket3-taxbracket2)*mrt2+income-taxbracket3*mrt3)


tax(10000)
tax(25000)
tax(35000)
tax(4000000)
